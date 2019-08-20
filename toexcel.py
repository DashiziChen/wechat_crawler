import pymongo
from openpyxl import Workbook
from openpyxl import load_workbook


def export2excel(gzh_list, filename):
    """
    将mongoDB的数据到出到excel
    :param gzh_list: 公众号列表，用于确定需要下载哪几张表的数据
    :param filename: 存储的excel文件名
    :return:
    """
    try:
        wb = load_workbook(filename)
    except:
        wb = Workbook()
    sheet = wb.active
    sheet.title = '微信采集'  # 对sheet进行命名
    sheet.cell(row=1, column=1).value = '新闻代码'
    sheet.cell(row=1, column=2).value = '新闻标题'
    sheet.cell(row=1, column=3).value = '发布时间'
    sheet.cell(row=1, column=4).value = '新闻作者'
    sheet.cell(row=1, column=5).value = '正文'
    sheet.cell(row=1, column=6).value = '新闻媒体'
    sheet.cell(row=1, column=7).value = '新闻网址'
    sheet.cell(row=1, column=8).value = '阅读量'
    sheet.cell(row=1, column=9).value = '评论量'
    sheet.cell(row=1, column=10).value = '转发量'
    sheet.cell(row=1, column=11).value = '点赞量'
    sheet.cell(row=1, column=12).value = '机构名称'

    myclient = pymongo.MongoClient("mongodb://localhost:27017/")
    mydb = myclient["wechat_crawler"]
    count = sheet.max_row + 1
    for gzh in gzh_list:
        mycol = mydb[gzh]
        for x in mycol.find():
            sheet.cell(row=count, column=2).value = x['title']
            sheet.cell(row=count, column=3).value = x['date'].split(" ")[0]
            sheet.cell(row=count, column=4).value = gzh
            sheet.cell(row=count, column=5).value = x['article']
            sheet.cell(row=count, column=6).value = "微信公众号"
            sheet.cell(row=count, column=7).value = x['url']
            sheet.cell(row=count, column=8).value = x['readNum']
            sheet.cell(row=count, column=9).value = x['comment_count']
            sheet.cell(row=count, column=11).value = x['likeNum']
            sheet.cell(row=count, column=12).value = gzh[0:2] + "银行股份有限公司"
            count += 1
    wb.save(filename)


if __name__ == '__main__':
    export2excel(["交通银行公司金融"])